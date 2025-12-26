"""
Doküman İşleme Modülü
PDF, DOCX ve XLSX dosyalarını okuyup işler.
"""

import os
from typing import List, Dict, Optional
import hashlib

try:
    from PyPDF2 import PdfReader
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


class DocumentProcessor:
    """PDF, DOCX ve XLSX dosyalarını işler"""
    
    def __init__(self, documents_folder: str = "documents"):
        self.documents_folder = documents_folder
        self.processed_files: Dict[str, str] = {}  # filename -> hash
        
        # Klasör yoksa oluştur
        if not os.path.exists(documents_folder):
            os.makedirs(documents_folder)
    
    def get_file_hash(self, filepath: str) -> str:
        """Dosyanın MD5 hash'ini hesaplar"""
        try:
            with open(filepath, 'rb') as f:
                return hashlib.md5(f.read()).hexdigest()
        except Exception:
            return ""
    
    def extract_text_from_pdf(self, filepath: str) -> str:
        """PDF dosyasından metin çıkarır"""
        if not PDF_AVAILABLE:
            return ""
        
        try:
            reader = PdfReader(filepath)
            text = ""
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n\n"
            return text.strip()
        except Exception as e:
            print(f"PDF okuma hatası ({filepath}): {e}")
            return ""
    
    def extract_text_from_docx(self, filepath: str) -> str:
        """DOCX dosyasından metin çıkarır"""
        if not DOCX_AVAILABLE:
            return ""
        
        try:
            doc = Document(filepath)
            text = ""
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    text += paragraph.text + "\n\n"
            return text.strip()
        except Exception as e:
            print(f"DOCX okuma hatası ({filepath}): {e}")
            return ""
    
    def extract_text_from_xlsx(self, filepath: str) -> str:
        """XLSX dosyasından metin çıkarır"""
        if not XLSX_AVAILABLE:
            return ""
        
        try:
            workbook = openpyxl.load_workbook(filepath)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"\n[Sayfa: {sheet_name}]\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        text += row_text + "\n"
            return text.strip()
        except Exception as e:
            print(f"XLSX okuma hatası ({filepath}): {e}")
            return ""
    
    def extract_text(self, filepath: str) -> str:
        """Dosya tipine göre metin çıkarır"""
        filename = os.path.basename(filepath).lower()
        
        if filename.endswith('.pdf'):
            return self.extract_text_from_pdf(filepath)
        elif filename.endswith('.docx'):
            return self.extract_text_from_docx(filepath)
        elif filename.endswith('.xlsx'):
            return self.extract_text_from_xlsx(filepath)
        else:
            return ""
    
    def get_all_documents(self) -> List[str]:
        """Dokümanlar klasöründeki tüm desteklenen dosyaları listeler"""
        documents = []
        if os.path.exists(self.documents_folder):
            for filename in os.listdir(self.documents_folder):
                filepath = os.path.join(self.documents_folder, filename)
                if os.path.isfile(filepath):
                    if filename.lower().endswith(('.pdf', '.docx', '.xlsx')):
                        documents.append(filepath)
        return documents
    
    def process_all_documents(self) -> List[Dict[str, str]]:
        """Tüm dokümanları işler ve metin listesi döndürür"""
        documents = []
        doc_files = self.get_all_documents()
        
        for filepath in doc_files:
            filename = os.path.basename(filepath)
            file_hash = self.get_file_hash(filepath)
            
            # Dosya zaten işlenmiş mi kontrol et
            if filename in self.processed_files and self.processed_files[filename] == file_hash:
                continue
            
            text = self.extract_text(filepath)
            if text:
                documents.append({
                    "filename": filename,
                    "content": text,
                    "filepath": filepath
                })
                self.processed_files[filename] = file_hash
        
        return documents
    
    def chunk_text(self, text: str, chunk_size: int = 1000, overlap: int = 200) -> List[str]:
        """Metni küçük parçalara böler"""
        chunks = []
        start = 0
        text_length = len(text)
        
        while start < text_length:
            end = start + chunk_size
            
            # Kelime ortasında bölmemeye çalış
            if end < text_length:
                last_space = text.rfind(' ', start, end)
                if last_space > start:
                    end = last_space
            
            chunks.append(text[start:end].strip())
            start = end - overlap
        
        return chunks
    
    def get_document_chunks(self) -> List[Dict[str, str]]:
        """Tüm dokümanları chunk'lara böler"""
        all_chunks = []
        documents = self.process_all_documents()
        
        for doc in documents:
            chunks = self.chunk_text(doc["content"])
            for i, chunk in enumerate(chunks):
                all_chunks.append({
                    "content": chunk,
                    "source": doc["filename"],
                    "chunk_id": f"{doc['filename']}_{i}"
                })
        
        return all_chunks


class SimpleDocumentStore:
    """Basit doküman deposu ve arama"""
    
    def __init__(self):
        self.documents: List[Dict[str, str]] = []
    
    def add_documents(self, chunks: List[Dict[str, str]]):
        """Doküman chunk'larını ekler"""
        self.documents.extend(chunks)
    
    def clear(self):
        """Tüm dokümanları temizler"""
        self.documents = []
    
    def search(self, query: str, top_k: int = 5) -> List[Dict[str, str]]:
        """Basit anahtar kelime araması yapar"""
        if not self.documents:
            return []
        
        query_words = set(query.lower().split())
        scored_docs = []
        
        for doc in self.documents:
            content_lower = doc["content"].lower()
            score = sum(1 for word in query_words if word in content_lower)
            if score > 0:
                scored_docs.append((score, doc))
        
        # Skora göre sırala
        scored_docs.sort(key=lambda x: x[0], reverse=True)
        
        return [doc for _, doc in scored_docs[:top_k]]
    
    def get_context(self, query: str, top_k: int = 3) -> str:
        """Sorgu için ilgili bağlamı döndürür"""
        results = self.search(query, top_k)
        if not results:
            return ""
        
        context_parts = []
        for doc in results:
            context_parts.append(f"[Kaynak: {doc['source']}]\n{doc['content']}")
        
        return "\n\n---\n\n".join(context_parts)
    
    def get_category_context(self, category: str) -> str:
        """Kategoriye göre ilgili bağlamı döndürür"""
        # Kategori anahtar kelimeleri
        category_keywords = {
            "motor": "motor motoru çalışmıyor marş ateşleme yakıt benzin dizel",
            "fren": "fren balata disk pedal sert ses",
            "elektrik": "akü batarya elektrik sigorta kablo alternatör",
            "klima": "klima ısıtma kalorifer soğutma fan",
            "sanziman": "şanzıman vites debriyaj aktarma",
            "bakim": "bakım yağ filtre servis muayene"
        }
        
        keywords = category_keywords.get(category.lower(), category)
        return self.get_context(keywords, top_k=5)

